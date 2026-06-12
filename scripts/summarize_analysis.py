#!/usr/bin/env python
"""
scripts/summarize_analysis.py
=============================

Collect every numeric result produced by ``run_analysis_all.py`` (and
its helpers) into a single compact *summary analysis table* that is easy
to read at a glance.

Motivation
----------
The pooled analysis already writes a large pile of per-condition artefacts
to the common analysis folder (``*_ps_metrics.json``, ``*_metrics.json``,
``*_msd_<pair>.json``, ``*_loop_fractions.json``, ``*_calibration.json``,
plus three condition-level tables: ``ps_derivative_table.json``,
``able_conserved_pairs_table.json``, ``msd_between_condition_stats.json``).

This script gathers the scalar summaries from every one of those files
and writes a single *metric-by-condition* matrix:

* Rows    = parameters (with a ``metric_group`` column for navigation)
* Columns = conditions (simulated + experimental, where applicable)

The matrix shape was chosen over "rows = conditions, columns = metrics"
because there are far more metrics (~30-40) than conditions (~2-5), so
the tall-narrow orientation reads like a parameter sheet and fits on
the screen without horizontal scrolling.

Outputs
-------
Given ``--out-prefix <common_dir>/summary_analysis_table``:

    summary_analysis_table.csv      pivoted CSV (metric_group, metric, <cond>...)
    summary_analysis_table.json     nested dict {group: {metric: {cond: value}}}
    summary_analysis_table.xlsx     optional, only if openpyxl is available

Significance testing (companion tables)
---------------------------------------
When at least two conditions have two or more successful per-replicate
fits, the same per-replicate JSONs that feed the pooled table are
harvested into a per-replicate matrix, and every (metric, pair of
conditions) is tested with a five-way battery: Welch's t, Mann-Whitney
U, label permutation (exact when feasible), Cohen's d, and a percentile
bootstrap 95% CI on the mean difference. K_alpha-like rows are compared
in log10 space automatically.

Two companion CSVs are written:

    summary_stats_all_pairs.csv     rows = metric, cols = "A vs B"
                                    for every ordered condition pair
    summary_stats_vs_reference.csv  rows = metric, cols = "<ref> vs B"
                                    against a single reference (default:
                                    first discovered condition; override
                                    with --reference-condition)

Plus a raw-values audit file:

    summary_stats_per_replicate.csv rows = metric, cols = one column per
                                    (condition, replicate) showing the
                                    input number that fed the tests

Pooled-only scalars (e.g. the P(s) derivative loop size, AbLE strength
on the pooled contact map) do not have a per-replicate distribution in
the current outputs; their cells in the two *_stats_* tables are left
explicitly empty rather than filled with a fabricated p-value.

Usage
-----
    python scripts/summarize_analysis.py \
        --analysis-dir results/polychrom_3d/../analysis

Or via the orchestrator (auto-wired as Phase 4):

    python scripts/run_analysis_all.py \
        --results-dir results/polychrom_3d \
        --mcool-mesc   data/hic/bonev_mESC.mcool \
        --mcool-neuron data/hic/bonev_CN.mcool

Pass ``--no-summary-table`` to the orchestrator to skip Phase 4.

Naming conventions we rely on
-----------------------------
The pooled files are written by ``run_analysis_all.py`` with this stem::

    <condition>_<nblocks>blk_pooled_<artefact>.{json,npz,png}

so ``<condition>`` is everything before ``_<digits>blk_pooled_``. We
discover pooled conditions by globbing ``*_pooled_ps_metrics.json``.

If a file for a given condition is missing, the corresponding cells are
left empty (not NaN-filled) so downstream viewers do not
accidentally treat "not computed" as "computed to be zero".
"""

from __future__ import annotations

import argparse
import glob
import itertools
import json
import logging
import os
import re
import sys
from collections import OrderedDict, defaultdict
from typing import Any, Dict, List, Optional, Tuple

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Regexes for filename discovery
# ---------------------------------------------------------------------------
#
#   <cond>_<n>blk_pooled_ps_metrics.json     <- condition anchor
#   <cond>_<n>blk_pooled_metrics.json        <- sim-vs-exp metrics
#   <cond>_<n>blk_pooled_msd_<label>.json    <- per-pair MSD
#   <cond>_<n>blk_pooled_loop_fractions.json
#   <cond>_<n>blk_pooled_calibration.json
#   <cond>_<n>blk_pooled_apa_loops_quant.json
#
# ``n`` is the number of pooled conformations and is kept for provenance
# but not emitted in the final summary table.

_POOLED_ANCHOR_RE = re.compile(
    r"^(?P<cond>.+)_(?P<nblocks>\d+)blk_pooled_ps_metrics\.json$")

_POOLED_MSD_RE = re.compile(
    r"^(?P<cond>.+)_(?P<nblocks>\d+)blk_pooled_msd_(?P<label>.+)\.json$")

# Per-replicate artefacts live next to the pooled files with filenames of the
# form ``<cond>_<n>blk_rep<R>_<suffix>.<ext>``. The rep token can be negative
# in smoke tests (``rep-1``), so we accept an optional minus sign.
_PER_REP_ANCHOR_RE = re.compile(
    r"^(?P<cond>.+)_(?P<nblocks>\d+)blk_rep(?P<rep>-?\d+)_(?P<suffix>.+)$"
)


# ---------------------------------------------------------------------------
# Metric groups (drives row ordering in the final table)
# ---------------------------------------------------------------------------
#
# The order here becomes the row order in the CSV. Within each group,
# individual metric names are appended in the order they first appear
# while scanning the files, and then sorted alphabetically inside the
# group to keep the output deterministic.
METRIC_GROUP_ORDER: List[str] = [
    "provenance",
    "calibration",
    "P(s)",
    "P(s)_derivative",
    "AbLE",
    "MSD",
    "dynamics",
    "sim_vs_exp",
]


# ===========================================================================
# CORE: discover conditions + harvest each metric family
# ===========================================================================

def discover_pooled_conditions(analysis_dir: str) -> "OrderedDict[str, int]":
    """Return ``{condition: nblocks}`` for every pooled condition found."""
    out: "OrderedDict[str, int]" = OrderedDict()
    for path in sorted(glob.glob(os.path.join(
            analysis_dir, "*_pooled_ps_metrics.json"))):
        m = _POOLED_ANCHOR_RE.match(os.path.basename(path))
        if not m:
            continue
        out[m.group("cond")] = int(m.group("nblocks"))
    return out


def _load_json(path: str) -> Optional[dict]:
    """Read a JSON file, returning None on any error (and logging at DEBUG)."""
    try:
        with open(path) as f:
            return json.load(f)
    except Exception as e:  # noqa: BLE001
        logger.debug(f"could not read {path}: {e}")
        return None


def _pooled_path(analysis_dir: str, cond: str, nblocks: int,
                 suffix: str) -> str:
    return os.path.join(
        analysis_dir, f"{cond}_{nblocks}blk_pooled_{suffix}")


def harvest_ps_metrics(analysis_dir: str, cond: str, nblocks: int,
                       out: Dict[str, Dict[str, Any]]) -> None:
    """Pull scalar P(s) metrics from ``<cond>_<n>blk_pooled_ps_metrics.json``."""
    p = _pooled_path(analysis_dir, cond, nblocks, "ps_metrics.json")
    payload = _load_json(p)
    if payload is None:
        return
    metrics = payload.get("metrics", {})
    fit = payload.get("fit", {})

    # P(s) magnitudes at reference separations
    for k, v in metrics.items():
        if not isinstance(v, (int, float)):
            continue
        # Keep anything that is not a derivative-at-X-kb (those belong
        # to the derivative group below)
        if k.startswith("dlogP_dlogs_at_"):
            out.setdefault("P(s)_derivative", {}).setdefault(k, {})[cond] = float(v)
        else:
            out.setdefault("P(s)", {}).setdefault(k, {})[cond] = float(v)

    # Power-law fit exponent / intercept
    for k in ("exponent", "log_intercept", "lower_bound_bp", "upper_bound_bp"):
        v = fit.get(k)
        if isinstance(v, (int, float)):
            out.setdefault("P(s)", {}).setdefault(f"fit_{k}", {})[cond] = float(v)


def harvest_calibration(analysis_dir: str, cond: str, nblocks: int,
                        out: Dict[str, Dict[str, Any]]) -> None:
    p = _pooled_path(analysis_dir, cond, nblocks, "calibration.json")
    payload = _load_json(p)
    if payload is None:
        return
    for k in ("nm_per_monomer", "sec_per_frame", "bp_per_monomer",
              "persistence_nm"):
        v = payload.get(k)
        if isinstance(v, (int, float)):
            out.setdefault("calibration", {}).setdefault(k, {})[cond] = float(v)
    src = payload.get("source")
    if isinstance(src, str):
        out.setdefault("calibration", {}).setdefault("source", {})[cond] = src


def harvest_sim_vs_exp(analysis_dir: str, cond: str, nblocks: int,
                       out: Dict[str, Dict[str, Any]]) -> None:
    p = _pooled_path(analysis_dir, cond, nblocks, "metrics.json")
    payload = _load_json(p)
    if payload is None:
        return
    for k in ("stratum_adjusted_corr", "overall_pearson",
              "per_stratum_mean_corr", "mse", "mae"):
        v = payload.get(k)
        if isinstance(v, (int, float)):
            out.setdefault("sim_vs_exp", {}).setdefault(k, {})[cond] = float(v)


def harvest_loop_fractions(analysis_dir: str, cond: str, nblocks: int,
                           out: Dict[str, Dict[str, Any]]) -> None:
    """Looped fraction per MSD pair label."""
    p = _pooled_path(analysis_dir, cond, nblocks, "loop_fractions.json")
    payload = _load_json(p)
    if not isinstance(payload, list):
        return
    for entry in payload:
        label = entry.get("label")
        frac = entry.get("looped_fraction")
        if label is None or frac is None:
            continue
        key = f"looped_fraction[{label}]"
        out.setdefault("dynamics", {}).setdefault(key, {})[cond] = float(frac)


def harvest_msd_pairs(analysis_dir: str, cond: str, nblocks: int,
                      out: Dict[str, Dict[str, Any]]) -> None:
    """Per-pair MSD fit scalars (alpha, K_alpha, saturation J / tau_c)."""
    pat = os.path.join(analysis_dir,
                       f"{cond}_{nblocks}blk_pooled_msd_*.json")
    for path in sorted(glob.glob(pat)):
        m = _POOLED_MSD_RE.match(os.path.basename(path))
        if not m:
            continue
        label = m.group("label")
        payload = _load_json(path)
        if payload is None:
            continue
        fit = payload.get("alpha_fit") or {}
        sat = payload.get("saturation_fit") or {}

        def _put(k: str, v: Any) -> None:
            if isinstance(v, (int, float)):
                key = f"{k}[{label}]"
                out.setdefault("MSD", {}).setdefault(key, {})[cond] = float(v)

        _put("alpha", fit.get("alpha"))
        _put("K_alpha", fit.get("K_alpha"))
        _put("alpha_stderr", fit.get("alpha_stderr"))
        _put("log_K_alpha_stderr", fit.get("log_K_alpha_stderr"))
        _put("sat_J", sat.get("J"))
        _put("sat_tau_c", sat.get("tau_c"))
        _put("sat_alpha", sat.get("alpha_sat"))
        # pair separation (monomers) for provenance
        pair = payload.get("pair", {})
        sep = pair.get("sep_monomers")
        if isinstance(sep, int):
            out.setdefault("MSD", {}).setdefault(
                f"sep_monomers[{label}]", {})[cond] = int(sep)


def harvest_apa_quant(analysis_dir: str, cond: str, nblocks: int,
                      out: Dict[str, Dict[str, Any]]) -> None:
    """Summarise batch_absolute_quant (number of loops + median strength)."""
    p = _pooled_path(analysis_dir, cond, nblocks, "apa_loops_quant.json")
    payload = _load_json(p)
    if payload is None:
        return
    loops = payload.get("loops") or []
    n_pairs = payload.get("n_pairs")
    if isinstance(n_pairs, int):
        out.setdefault("AbLE", {}).setdefault("apa_n_pairs", {})[cond] = int(n_pairs)
    if loops:
        strengths = [l.get("strength") for l in loops
                     if isinstance(l.get("strength"), (int, float))]
        if strengths:
            strengths_sorted = sorted(strengths)
            n = len(strengths_sorted)
            median = (strengths_sorted[n // 2]
                      if n % 2 == 1
                      else 0.5 * (strengths_sorted[n // 2 - 1]
                                  + strengths_sorted[n // 2]))
            mean = sum(strengths_sorted) / n
            out.setdefault("AbLE", {}).setdefault(
                "apa_strength_median", {})[cond] = float(median)
            out.setdefault("AbLE", {}).setdefault(
                "apa_strength_mean", {})[cond] = float(mean)


def harvest_rg(analysis_dir: str, cond: str, nblocks: int,
               out: Dict[str, Dict[str, Any]]) -> None:
    """Radius of gyration: mean + std over the timecourse."""
    # We wrote rg_timecourse.npz; the JSON summary is stored inside
    # the per-pair MSD json's extra dict. Load the npz directly when
    # available (lazy import numpy so the script still works without it
    # when only JSON files are present).
    p = _pooled_path(analysis_dir, cond, nblocks, "rg_timecourse.npz")
    if not os.path.exists(p):
        return
    try:
        import numpy as _np
        arr = _np.load(p)
        rg = arr.get("rg") if hasattr(arr, "get") else arr["rg"]
        if rg is None or rg.size == 0:
            return
        out.setdefault("dynamics", {}).setdefault(
            "rg_mean", {})[cond] = float(rg.mean())
        out.setdefault("dynamics", {}).setdefault(
            "rg_std", {})[cond] = float(rg.std())
    except Exception as e:  # noqa: BLE001
        logger.debug(f"could not read {p}: {e}")


# ===========================================================================
# CONDITION-LEVEL TABLES (already aggregated across conditions)
# ===========================================================================
# These tables live at the top level of ``common_dir`` and already contain
# a row per condition. We flatten them into the same metric-by-condition
# matrix so everything ends up in one place.

def harvest_ps_derivative_table(
    analysis_dir: str,
    out: Dict[str, Dict[str, Any]],
) -> List[str]:
    """
    Read ``ps_derivative_table.json`` and flatten into the output dict.

    Returns the list of condition labels seen there (sim + experimental).
    Experimental rows have labels like ``mESC_experimental (exp)`` and
    are kept in the table so sim-vs-in-vivo comparison is one column away.
    """
    p = os.path.join(analysis_dir, "ps_derivative_table.json")
    payload = _load_json(p)
    if payload is None:
        return []
    table = payload.get("table") or {}
    for cond, row in table.items():
        for metric, value in row.items():
            if isinstance(value, (int, float)):
                out.setdefault("P(s)_derivative", {}).setdefault(
                    metric, {})[cond] = float(value)
            elif (isinstance(value, list)
                  and all(isinstance(v, (int, float)) for v in value)
                  and len(value) == 2):
                # e.g. loop_size_search_window_bp = [lo, hi] - keep as "lo-hi"
                out.setdefault("P(s)_derivative", {}).setdefault(
                    metric, {})[cond] = f"{value[0]:g}-{value[1]:g}"
    return list(table.keys())


def harvest_able_table(
    analysis_dir: str,
    out: Dict[str, Dict[str, Any]],
) -> List[str]:
    p = os.path.join(analysis_dir, "able_conserved_pairs_table.json")
    payload = _load_json(p)
    if payload is None:
        return []
    pair_labels = payload.get("pair_labels") or []
    strength = payload.get("strength") or {}
    center = payload.get("center_mean") or {}
    bkg = payload.get("background_mean") or {}
    conditions_seen: set = set()
    for plabel in pair_labels:
        s_map = strength.get(plabel, {}) or {}
        c_map = center.get(plabel, {}) or {}
        b_map = bkg.get(plabel, {}) or {}
        for cond, v in s_map.items():
            if isinstance(v, (int, float)):
                out.setdefault("AbLE", {}).setdefault(
                    f"able_strength[{plabel}]", {})[cond] = float(v)
                conditions_seen.add(cond)
        for cond, v in c_map.items():
            if isinstance(v, (int, float)):
                out.setdefault("AbLE", {}).setdefault(
                    f"able_center[{plabel}]", {})[cond] = float(v)
        for cond, v in b_map.items():
            if isinstance(v, (int, float)):
                out.setdefault("AbLE", {}).setdefault(
                    f"able_background[{plabel}]", {})[cond] = float(v)
    return sorted(conditions_seen)


def harvest_msd_stats(
    analysis_dir: str,
    out: Dict[str, Dict[str, Any]],
) -> List[str]:
    """
    Flatten ``msd_between_condition_stats.json`` per-condition summaries.

    This file stores per-replicate alpha / K_alpha and bootstrap summary
    dicts; we emit the summary means for every (pair, condition) cell so
    they live next to the pooled single-fit alpha / K_alpha from the
    per-pair MSD JSONs (which come from ``harvest_msd_pairs`` above).
    """
    p = os.path.join(analysis_dir, "msd_between_condition_stats.json")
    payload = _load_json(p)
    if payload is None:
        return []
    pairs = payload.get("pairs") or {}
    conditions_seen: set = set()
    for pair_label, pair_entry in pairs.items():
        for cond, entry in (pair_entry.get("conditions") or {}).items():
            a_sum = entry.get("alpha_summary") or {}
            k_sum = entry.get("K_alpha_summary") or {}
            logk_sum = entry.get("log10_K_alpha_summary") or {}
            for metric_key, summary in (
                (f"alpha_mean_perrep[{pair_label}]", a_sum),
                (f"K_alpha_mean_perrep[{pair_label}]", k_sum),
                (f"log10_K_alpha_mean_perrep[{pair_label}]", logk_sum),
            ):
                mean = summary.get("mean") if isinstance(summary, dict) else None
                sem = summary.get("sem") if isinstance(summary, dict) else None
                if isinstance(mean, (int, float)):
                    out.setdefault("MSD", {}).setdefault(
                        metric_key, {})[cond] = float(mean)
                    conditions_seen.add(cond)
                if isinstance(sem, (int, float)):
                    out.setdefault("MSD", {}).setdefault(
                        metric_key.replace("_mean_perrep[",
                                           "_sem_perrep["),
                        {})[cond] = float(sem)
            n = entry.get("alpha_summary", {}).get("n")
            if isinstance(n, int):
                out.setdefault("MSD", {}).setdefault(
                    f"n_replicates_fit[{pair_label}]", {})[cond] = int(n)
                conditions_seen.add(cond)
    return sorted(conditions_seen)


# ===========================================================================
# PER-REPLICATE HARVEST (for significance testing)
# ===========================================================================
#
# The pooled tables above are single-number-per-condition. To get p-values we
# need the per-replicate vectors that were averaged / pooled into those
# numbers. We harvest them from the same common analysis directory by
# globbing ``*_rep<N>_*`` filenames and extracting the same scalars the
# pooled harvesters already extract, keeping the metric names identical so
# the two tables line up row-for-row.
#
# Output structure is nested three deep::
#
#     {metric_group: {metric_name: {condition: [v_rep1, v_rep2, ...]}}}

def _per_rep_path(analysis_dir: str, cond: str, nblocks: int, rep: int,
                  suffix: str) -> str:
    return os.path.join(
        analysis_dir, f"{cond}_{nblocks}blk_rep{rep}_{suffix}")


def discover_per_replicate(
    analysis_dir: str,
) -> Dict[str, Dict[int, List[int]]]:
    """
    Return ``{condition: {nblocks: sorted list of rep numbers}}`` for every
    per-replicate artefact found in ``analysis_dir``.

    We key by both condition and nblocks because nblocks is part of the
    filename convention (it records the number of conformation blocks per
    replicate); in practice it is constant within a condition but harvesting
    is defensive.
    """
    index: Dict[str, Dict[int, set]] = defaultdict(lambda: defaultdict(set))
    for path in sorted(glob.glob(os.path.join(analysis_dir, "*_rep*_*"))):
        name = os.path.basename(path)
        m = _PER_REP_ANCHOR_RE.match(name)
        if not m:
            continue
        cond = m.group("cond")
        try:
            nblocks = int(m.group("nblocks"))
            rep = int(m.group("rep"))
        except ValueError:
            continue
        index[cond][nblocks].add(rep)
    return {c: {n: sorted(s) for n, s in nb.items()}
            for c, nb in index.items()}


def harvest_per_replicate_scalars(
    analysis_dir: str,
) -> Tuple[Dict[str, Dict[str, Dict[str, List[float]]]],
           Dict[str, Dict[str, Dict[str, Dict[Tuple[int, int], float]]]],
           Dict[str, List[Tuple[int, int]]]]:
    """
    Walk every per-replicate JSON in ``analysis_dir`` and collect the same
    scalar metrics the pooled harvesters collect. For each (metric_group,
    metric_name, condition) cell we accumulate a list of per-rep values in
    ascending rep-number order.

    Returns
    -------
    per_rep : dict
        Nested ``{metric_group: {metric_name: {condition: [val_rep_a, ...]}}}``
        (values only, skipping any replicate that did not contribute). This
        is what the pairwise stats runner consumes.
    per_rep_by_rep : dict
        Parallel structure
        ``{metric_group: {metric_name: {condition: {(nblocks, rep): value}}}}``
        that preserves the mapping from (nblocks, rep) to value. Used by
        the audit CSV so each replicate's contribution is kept on its own
        column even when some metrics are missing for some reps.
    reps_per_cond : dict
        ``{condition: [(nblocks, rep), ...]}`` listing every (nblocks, rep)
        tuple that was actually read for that condition.
    """
    per_rep: Dict[str, Dict[str, Dict[str, List[float]]]] = {}
    per_rep_by_rep: Dict[
        str, Dict[str, Dict[str, Dict[Tuple[int, int], float]]]] = {}
    reps_per_cond: Dict[str, List[Tuple[int, int]]] = defaultdict(list)

    def _put(group: str, metric: str, cond: str,
             nblocks: int, rep: int, value: float) -> None:
        per_rep.setdefault(group, {}).setdefault(
            metric, {}).setdefault(cond, []).append(float(value))
        per_rep_by_rep.setdefault(group, {}).setdefault(
            metric, {}).setdefault(cond, {})[(int(nblocks), int(rep))] = float(value)

    index = discover_per_replicate(analysis_dir)

    # Iterate in a deterministic order so the rep list matches across metrics
    # for the same (condition, rep). Within a condition we walk reps in
    # ascending order; one (nblocks, rep) tuple at a time.
    for cond in sorted(index):
        for nblocks in sorted(index[cond]):
            for rep in index[cond][nblocks]:
                reps_per_cond[cond].append((nblocks, rep))

                # ---- P(s) scalars ----
                p = _per_rep_path(analysis_dir, cond, nblocks, rep,
                                  "ps_metrics.json")
                payload = _load_json(p)
                if payload is not None:
                    metrics = payload.get("metrics", {}) or {}
                    fit = payload.get("fit", {}) or {}
                    for k, v in metrics.items():
                        if not isinstance(v, (int, float)):
                            continue
                        if k.startswith("dlogP_dlogs_at_"):
                            _put("P(s)_derivative", k, cond, nblocks, rep, v)
                        else:
                            _put("P(s)", k, cond, nblocks, rep, v)
                    for k in ("exponent", "log_intercept",
                              "lower_bound_bp", "upper_bound_bp"):
                        v = fit.get(k)
                        if isinstance(v, (int, float)):
                            _put("P(s)", f"fit_{k}", cond, nblocks, rep, v)

                # ---- sim-vs-exp scalars ----
                p = _per_rep_path(analysis_dir, cond, nblocks, rep,
                                  "metrics.json")
                payload = _load_json(p)
                if payload is not None:
                    for k in ("stratum_adjusted_corr", "overall_pearson",
                              "per_stratum_mean_corr", "mse", "mae"):
                        v = payload.get(k)
                        if isinstance(v, (int, float)):
                            _put("sim_vs_exp", k, cond, nblocks, rep, v)

                # ---- looped fractions ----
                p = _per_rep_path(analysis_dir, cond, nblocks, rep,
                                  "loop_fractions.json")
                payload = _load_json(p)
                if isinstance(payload, list):
                    for entry in payload:
                        label = entry.get("label")
                        frac = entry.get("looped_fraction")
                        if label is None or frac is None:
                            continue
                        _put("dynamics",
                             f"looped_fraction[{label}]", cond,
                             nblocks, rep, frac)

                # ---- per-pair MSD fits ----
                pat = os.path.join(
                    analysis_dir,
                    f"{cond}_{nblocks}blk_rep{rep}_msd_*.json")
                for msd_path in sorted(glob.glob(pat)):
                    mm = re.match(
                        rf"^{re.escape(cond)}_{nblocks}blk_rep{rep}_msd_"
                        r"(?P<label>.+)\.json$",
                        os.path.basename(msd_path))
                    if not mm:
                        continue
                    label = mm.group("label")
                    payload = _load_json(msd_path)
                    if payload is None:
                        continue
                    fit = payload.get("alpha_fit") or {}
                    sat = payload.get("saturation_fit") or {}
                    for k_json, k_out in (
                        ("alpha", "alpha"),
                        ("K_alpha", "K_alpha"),
                        ("alpha_stderr", "alpha_stderr"),
                        ("log_K_alpha_stderr", "log_K_alpha_stderr"),
                    ):
                        v = fit.get(k_json)
                        if isinstance(v, (int, float)):
                            _put("MSD", f"{k_out}[{label}]", cond,
                                 nblocks, rep, v)
                    for k_json, k_out in (
                        ("J", "sat_J"),
                        ("tau_c", "sat_tau_c"),
                        ("alpha_sat", "sat_alpha"),
                    ):
                        v = sat.get(k_json)
                        if isinstance(v, (int, float)):
                            _put("MSD", f"{k_out}[{label}]", cond,
                                 nblocks, rep, v)

                # ---- calibration scalars ----
                p = _per_rep_path(analysis_dir, cond, nblocks, rep,
                                  "calibration.json")
                payload = _load_json(p)
                if payload is not None:
                    for k in ("nm_per_monomer", "sec_per_frame",
                              "bp_per_monomer", "persistence_nm"):
                        v = payload.get(k)
                        if isinstance(v, (int, float)):
                            _put("calibration", k, cond, nblocks, rep, v)

                # ---- APA / AbLE strength scalars ----
                p = _per_rep_path(analysis_dir, cond, nblocks, rep,
                                  "apa_loops_quant.json")
                payload = _load_json(p)
                if payload is not None:
                    loops = payload.get("loops") or []
                    n_pairs = payload.get("n_pairs")
                    if isinstance(n_pairs, int):
                        _put("AbLE", "apa_n_pairs", cond,
                             nblocks, rep, float(n_pairs))
                    strengths = [l.get("strength") for l in loops
                                 if isinstance(l.get("strength"),
                                               (int, float))]
                    if strengths:
                        ss = sorted(strengths)
                        n = len(ss)
                        median = (ss[n // 2] if n % 2 == 1
                                  else 0.5 * (ss[n // 2 - 1] + ss[n // 2]))
                        mean = sum(ss) / n
                        _put("AbLE", "apa_strength_median", cond,
                             nblocks, rep, median)
                        _put("AbLE", "apa_strength_mean", cond,
                             nblocks, rep, mean)

    return per_rep, per_rep_by_rep, dict(reps_per_cond)


# ===========================================================================
# PAIRWISE SIGNIFICANCE TESTING
# ===========================================================================

# Metrics where the biologically interesting scale is multiplicative rather
# than additive. These are compared in log10 space (matches the MSD module's
# convention for K_alpha). The test here is substring-based so it also
# catches the per-pair variants like ``K_alpha[conv_100kb]``.
_LOG10_METRICS_SUBSTR = ("K_alpha", "looped_fraction",
                         "apa_strength_mean", "apa_strength_median",
                         "apa_n_pairs")


def _should_log_transform(metric_name: str) -> bool:
    return any(s in metric_name for s in _LOG10_METRICS_SUBSTR)


def _format_stats_cell(result: Dict[str, Any]) -> str:
    """
    Compact single-cell string summarising a comparison. Empty if the
    inputs did not have enough data for a meaningful test.
    """
    sx = result.get("summary_x") or {}
    sy = result.get("summary_y") or {}
    nx, ny = sx.get("n", 0), sy.get("n", 0)
    if nx == 0 or ny == 0:
        return ""
    diff = result.get("diff_mean", float("nan"))
    ci = result.get("diff_ci95") or {}
    lo = ci.get("lo", float("nan"))
    hi = ci.get("hi", float("nan"))
    d = result.get("cohens_d", float("nan"))
    pw = (result.get("welch_t") or {}).get("p_value", float("nan"))
    pmw = (result.get("mann_whitney") or {}).get("p_value", float("nan"))
    pp = (result.get("permutation") or {}).get("p_value", float("nan"))

    def _fmt(x: float) -> str:
        try:
            x = float(x)
        except (TypeError, ValueError):
            return "nan"
        if x != x:  # NaN
            return "nan"
        return f"{x:.3g}"

    tag = " [log10]" if result.get("log10_transformed") else ""
    return (
        f"{_fmt(diff)}{tag} [{_fmt(lo)}, {_fmt(hi)}], "
        f"d={_fmt(d)}, "
        f"p_w={_fmt(pw)}/p_mw={_fmt(pmw)}/p_perm={_fmt(pp)}, "
        f"n={nx} vs {ny}"
    )


def run_pairwise_stats(
    per_rep: Dict[str, Dict[str, Dict[str, List[float]]]],
    conditions: List[str],
    pair_mode: str,
    reference: Optional[str] = None,
    n_boot: int = 10_000,
    seed: int = 0,
    min_n_per_condition: int = 2,
) -> Tuple[Dict[str, Dict[str, Dict[str, Dict[str, Any]]]], List[str]]:
    """
    Run the between-condition battery for every (metric, condition pair).

    Parameters
    ----------
    per_rep : dict
        Output of ``harvest_per_replicate_scalars`` (nested three deep).
    conditions : list of str
        Condition order to use for pair enumeration. Only conditions that
        are present in ``per_rep`` for a given metric are actually compared.
    pair_mode : {"all_ordered", "vs_reference"}
        "all_ordered" compares every ordered pair (A, B) with A != B.
        "vs_reference" compares the reference condition against every other
        condition.
    reference : str, optional
        Required when ``pair_mode == "vs_reference"``. Must be in
        ``conditions``.
    n_boot : int
        Bootstrap resamples for the CI on the difference of means.
    seed : int
        Seed for the bootstrap / permutation RNGs.
    min_n_per_condition : int
        Minimum successful per-rep values a condition must contribute to
        participate in a test. Cells where either side falls below this
        threshold are left empty.

    Returns
    -------
    stats_nested : dict
        ``{metric_group: {metric_name: {pair_label: result_dict}}}`` where
        ``result_dict`` is the full output of ``compare_vectors``.
    pair_labels : list of str
        Column order for the output CSV (preserves the condition order
        passed in).
    """
    # Late import so the plain harvest path still works on machines where
    # analysis/msd_statistics.py has an import-time issue.
    analysis_dir = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "analysis")
    sys.path.insert(0, os.path.abspath(analysis_dir))
    try:
        from msd_statistics import compare_vectors  # noqa: WPS433
    except Exception as e:  # noqa: BLE001
        logger.warning(
            f"  could not import msd_statistics.compare_vectors ({e}); "
            f"skipping significance testing.")
        return {}, []

    if pair_mode == "vs_reference":
        if reference is None or reference not in conditions:
            logger.warning(
                f"  vs_reference pair_mode requires a reference in "
                f"{conditions!r}; got {reference!r}. Skipping.")
            return {}, []
        pairs = [(reference, c) for c in conditions if c != reference]
    elif pair_mode == "all_ordered":
        pairs = [(a, b) for a in conditions for b in conditions if a != b]
    else:
        raise ValueError(f"unknown pair_mode {pair_mode!r}")

    pair_labels = [f"{a} vs {b}" for (a, b) in pairs]
    stats_nested: Dict[str, Dict[str, Dict[str, Dict[str, Any]]]] = {}

    for group, metrics in per_rep.items():
        for metric_name, cond_map in metrics.items():
            log_tf = _should_log_transform(metric_name)
            for (a, b) in pairs:
                x = cond_map.get(a) or []
                y = cond_map.get(b) or []
                x = [v for v in x if isinstance(v, (int, float))
                     and v == v]  # drop NaN
                y = [v for v in y if isinstance(v, (int, float))
                     and v == v]
                label = f"{a} vs {b}"
                # Leave cell empty if either side is below the threshold.
                # compare_vectors would still return a dict but the test
                # p-values would be NaN; formatting them would be noise.
                if len(x) < min_n_per_condition or len(y) < min_n_per_condition:
                    continue
                try:
                    result = compare_vectors(
                        x, y, name_x=a, name_y=b,
                        log_transform=log_tf,
                        n_boot=n_boot, seed=seed,
                    )
                except Exception as e:  # noqa: BLE001
                    logger.debug(
                        f"compare_vectors failed on "
                        f"({group}/{metric_name}, {a} vs {b}): {e}")
                    continue
                stats_nested.setdefault(group, {}).setdefault(
                    metric_name, {})[label] = result

    return stats_nested, pair_labels


# -- stats writers ---------------------------------------------------------

def write_stats_csv(
    stats_nested: Dict[str, Dict[str, Dict[str, Dict[str, Any]]]],
    pair_labels: List[str],
    out_path: str,
) -> None:
    """Pivot stats into a metric-by-pair CSV (same row order as summary)."""
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    groups = _sorted_metric_groups(stats_nested)
    header = ["metric_group", "metric"] + pair_labels
    with open(out_path, "w") as f:
        f.write(",".join(header) + "\n")
        for g in groups:
            metrics = sorted(stats_nested[g].keys())
            for m in metrics:
                pair_map = stats_nested[g][m]
                cells = [g, m] + [
                    _format_stats_cell(pair_map.get(p, {}))
                    for p in pair_labels
                ]
                safe = [
                    '"' + x.replace('"', '""') + '"' if ("," in x or '"' in x)
                    else x
                    for x in cells
                ]
                f.write(",".join(safe) + "\n")
    logger.info(f"  stats CSV    → {out_path}")


def write_stats_json(
    stats_nested: Dict[str, Dict[str, Dict[str, Dict[str, Any]]]],
    pair_labels: List[str],
    pair_mode: str,
    reference: Optional[str],
    n_boot: int,
    out_path: str,
) -> None:
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    payload = {
        "schema_version": 1,
        "pair_mode": pair_mode,
        "reference": reference,
        "pair_labels": pair_labels,
        "n_boot": int(n_boot),
        "caveats": [
            "Each cell is the full output of msd_statistics.compare_vectors: "
            "Welch's t, Mann-Whitney U, permutation, Cohen's d, and a "
            "percentile bootstrap 95 percent CI on the mean difference.",
            "K_alpha-like metrics (K_alpha, looped_fraction, apa_strength) "
            "are compared in log10 space; diff_mean and CI are in log10 "
            "units for those rows.",
            "With n_rep = 3 per condition, the smallest two-sided exact "
            "permutation p is 2/20 = 0.10. Use bootstrap CI and Cohen's d "
            "as the primary interpretation anchors at small n.",
            "Empty cells mean one side had fewer than min_n_per_condition "
            "values; no test was run.",
        ],
        "stats": stats_nested,
    }
    with open(out_path, "w") as f:
        json.dump(payload, f, indent=2, default=float)
    logger.info(f"  stats JSON   → {out_path}")


def write_per_replicate_csv(
    per_rep_by_rep: Dict[str, Dict[str, Dict[str, Dict[Tuple[int, int], float]]]],
    reps_per_cond: Dict[str, List[Tuple[int, int]]],
    out_path: str,
) -> None:
    """
    Audit CSV: one column per (condition, rep) giving the raw value that
    fed the tests. Rows = metrics, same order as the stats CSV. Lookups
    use ``per_rep_by_rep`` so missing replicate values leave an empty cell
    rather than shifting other values into the wrong column.
    """
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    # Build the column order once: sort conditions alphabetically for
    # reproducibility; reps in ascending order within a condition.
    col_tuples: List[Tuple[str, int, int]] = []
    for cond in sorted(reps_per_cond):
        for nblocks, rep in reps_per_cond[cond]:
            col_tuples.append((cond, nblocks, rep))
    col_headers = [f"{c}_rep{r}" for (c, _n, r) in col_tuples]

    groups = _sorted_metric_groups(per_rep_by_rep)
    header = ["metric_group", "metric"] + col_headers
    with open(out_path, "w") as f:
        f.write(",".join(header) + "\n")
        for g in groups:
            metrics = sorted(per_rep_by_rep[g].keys())
            for m in metrics:
                cond_map = per_rep_by_rep[g][m]
                row: List[str] = [g, m]
                for (cond, nblocks, rep) in col_tuples:
                    val = (cond_map.get(cond) or {}).get((nblocks, rep))
                    if val is None:
                        row.append("")
                    else:
                        row.append(_format_cell(val))
                safe = [
                    '"' + x.replace('"', '""') + '"' if ("," in x or '"' in x)
                    else x
                    for x in row
                ]
                f.write(",".join(safe) + "\n")
    logger.info(f"  per-rep CSV  → {out_path}")


# ===========================================================================
# ASSEMBLE + WRITE
# ===========================================================================

def _union_ordered(*lists: List[str]) -> List[str]:
    """Union of multiple lists, preserving first-seen order."""
    out: List[str] = []
    seen: set = set()
    for lst in lists:
        for x in lst:
            if x not in seen:
                out.append(x)
                seen.add(x)
    return out


def collect_summary(analysis_dir: str) -> Tuple[
        Dict[str, Dict[str, Dict[str, Any]]], List[str], Dict[str, int]]:
    """
    Walk ``analysis_dir`` and build the nested dict
    ``{metric_group: {metric_name: {condition: value}}}``.

    Returns ``(nested, condition_order, condition_nblocks)`` where
    ``condition_order`` is the final ordered column list (sim conditions
    first, experimental conditions last) and ``condition_nblocks`` is a
    ``{sim_cond: nblocks}`` map (kept for provenance, not emitted).
    """
    if not os.path.isdir(analysis_dir):
        raise FileNotFoundError(analysis_dir)

    nested: Dict[str, Dict[str, Dict[str, Any]]] = {}

    # --- per-condition files (one set per pooled simulated condition) ---
    sim_conds = discover_pooled_conditions(analysis_dir)
    for cond, nblocks in sim_conds.items():
        harvest_ps_metrics(analysis_dir, cond, nblocks, nested)
        harvest_calibration(analysis_dir, cond, nblocks, nested)
        harvest_sim_vs_exp(analysis_dir, cond, nblocks, nested)
        harvest_loop_fractions(analysis_dir, cond, nblocks, nested)
        harvest_msd_pairs(analysis_dir, cond, nblocks, nested)
        harvest_apa_quant(analysis_dir, cond, nblocks, nested)
        harvest_rg(analysis_dir, cond, nblocks, nested)
        # record the number of pooled conformations
        nested.setdefault("provenance", {}).setdefault(
            "n_pooled_conformations", {})[cond] = int(nblocks)

    # --- condition-level tables (include experimental rows if present) ---
    deriv_conds = harvest_ps_derivative_table(analysis_dir, nested)
    able_conds = harvest_able_table(analysis_dir, nested)
    stat_conds = harvest_msd_stats(analysis_dir, nested)

    # Column order: sim conditions first (in discovery order),
    # then anything new seen only in derivative/AbLE/stats tables
    # (typically the experimental rows), preserving the order they
    # appeared in their source files.
    sim_order = list(sim_conds.keys())
    other_order = _union_ordered(deriv_conds, able_conds, stat_conds)
    extras = [c for c in other_order if c not in sim_conds]
    cond_order = sim_order + extras

    return nested, cond_order, dict(sim_conds)


# -- writers ----------------------------------------------------------------

def _sorted_metric_groups(
    nested: Dict[str, Dict[str, Dict[str, Any]]],
) -> List[str]:
    """Return groups in preferred order, with any unexpected ones appended."""
    known = [g for g in METRIC_GROUP_ORDER if g in nested]
    extra = sorted(g for g in nested if g not in METRIC_GROUP_ORDER)
    return known + extra


def _format_cell(v: Any) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float):
        if abs(v) >= 1e4 or (0 < abs(v) < 1e-3):
            return f"{v:.4g}"
        return f"{v:.4f}"
    if isinstance(v, int):
        return str(v)
    return str(v)


def write_csv(
    nested: Dict[str, Dict[str, Dict[str, Any]]],
    cond_order: List[str],
    out_path: str,
) -> None:
    """Metric-by-condition CSV with a leading ``metric_group`` column."""
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    groups = _sorted_metric_groups(nested)
    header = ["metric_group", "metric"] + cond_order
    with open(out_path, "w") as f:
        f.write(",".join(header) + "\n")
        for g in groups:
            metrics = sorted(nested[g].keys())
            for m in metrics:
                row_map = nested[g][m]
                cells = [g, m] + [_format_cell(row_map.get(c, ""))
                                  for c in cond_order]
                # Escape commas / quotes the standard CSV way
                safe = [
                    '"' + x.replace('"', '""') + '"' if ("," in x or '"' in x)
                    else x
                    for x in cells
                ]
                f.write(",".join(safe) + "\n")
    logger.info(f"  summary CSV  → {out_path}")


def write_json(
    nested: Dict[str, Dict[str, Dict[str, Any]]],
    cond_order: List[str],
    condition_nblocks: Dict[str, int],
    out_path: str,
) -> None:
    """Nested JSON with the same content as the CSV, plus metadata header."""
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    groups = _sorted_metric_groups(nested)
    clean: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for g in groups:
        clean[g] = {m: dict(nested[g][m]) for m in sorted(nested[g].keys())}
    payload = {
        "schema_version": 1,
        "description": (
            "Pivoted summary analysis table. Rows are metrics (grouped "
            "by metric_group); columns are conditions. See cohesin_sim/"
            "README.md section 8 for the meaning of each metric."
        ),
        "condition_order": cond_order,
        "condition_nblocks": condition_nblocks,
        "metric_group_order": groups,
        "table": clean,
    }
    with open(out_path, "w") as f:
        json.dump(payload, f, indent=2)
    logger.info(f"  summary JSON → {out_path}")


def write_xlsx(
    nested: Dict[str, Dict[str, Dict[str, Any]]],
    cond_order: List[str],
    out_path: str,
) -> None:
    """
    Optional XLSX output. Requires openpyxl. Bold header + group column +
    frozen top row for easier reading in Excel / LibreOffice.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.info("  (openpyxl not available; skipping XLSX)")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    groups = _sorted_metric_groups(nested)
    header = ["metric_group", "metric"] + cond_order
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    group_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    for g in groups:
        metrics = sorted(nested[g].keys())
        for m in metrics:
            row_map = nested[g][m]
            row: List[Any] = [g, m]
            for c in cond_order:
                v = row_map.get(c, "")
                if isinstance(v, (int, float)):
                    row.append(v)
                else:
                    row.append(_format_cell(v))
            ws.append(row)
            # mild visual grouping: shade the metric_group cell
            ws.cell(row=ws.max_row, column=1).fill = group_fill

    ws.freeze_panes = "C2"
    # autosize-ish
    for col_idx, col_name in enumerate(header, start=1):
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = max(14, min(30, len(col_name) + 4))

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    logger.info(f"  summary XLSX → {out_path}")


# ===========================================================================
# PUBLIC ENTRY POINT
# ===========================================================================

def build_summary_table(
    analysis_dir: str,
    out_prefix: Optional[str] = None,
    write_xlsx_file: bool = True,
    do_stats: bool = True,
    reference_condition: Optional[str] = None,
    n_boot: int = 10_000,
    stats_seed: int = 0,
    min_n_per_condition: int = 2,
) -> Tuple[str, str, Optional[str]]:
    """
    End-to-end: discover, harvest, write CSV + JSON (+ XLSX if possible),
    plus two companion significance-test CSVs and a per-replicate audit
    CSV when ``do_stats`` is True and enough replicates are present.

    Parameters
    ----------
    analysis_dir : str
        Directory where ``run_analysis_all.py`` wrote the pooled outputs.
    out_prefix : str, optional
        Path prefix (no extension) for the main summary outputs. Defaults
        to ``<analysis_dir>/summary_analysis_table``. The stats CSVs are
        written next to the main summary as ``summary_stats_all_pairs``,
        ``summary_stats_vs_reference``, and ``summary_stats_per_replicate``.
    write_xlsx_file : bool
        Write XLSX if openpyxl is importable. No-op otherwise.
    do_stats : bool
        If True, harvest the per-replicate JSONs, run the comparison
        battery, and write the two stats CSVs plus the raw per-rep CSV.
    reference_condition : str, optional
        Condition to use as the reference column in the vs-reference CSV.
        Defaults to the first condition discovered (alphabetical within
        the pooled set; matches ``condition_order``).
    n_boot : int
        Bootstrap / permutation iterations. Default 10 000.
    stats_seed : int
        Seed for the RNG used in bootstrap and permutation.
    min_n_per_condition : int
        Minimum successful per-rep values a condition must contribute to
        participate in a test. Cells below the threshold are left empty.

    Returns
    -------
    (csv_path, json_path, xlsx_path_or_None)
        Paths of the primary summary outputs. Stats CSVs are logged but
        not returned (they sit next to the main outputs).
    """
    analysis_dir = os.path.abspath(analysis_dir)
    if out_prefix is None:
        out_prefix = os.path.join(analysis_dir, "summary_analysis_table")

    nested, cond_order, nblocks_map = collect_summary(analysis_dir)
    if not cond_order:
        logger.warning(
            f"  no pooled conditions or condition-level tables found in "
            f"{analysis_dir}; skipping summary.")
        return "", "", None
    if not nested:
        logger.warning(
            f"  found {len(cond_order)} conditions but no scalar metrics "
            f"in {analysis_dir}; nothing to summarise.")
        return "", "", None

    logger.info(
        f"  summarising {sum(len(m) for m in nested.values())} metrics "
        f"across {len(cond_order)} conditions "
        f"({', '.join(cond_order)})")

    csv_path = out_prefix + ".csv"
    json_path = out_prefix + ".json"
    xlsx_path: Optional[str] = out_prefix + ".xlsx" if write_xlsx_file else None

    write_csv(nested, cond_order, csv_path)
    write_json(nested, cond_order, nblocks_map, json_path)
    if write_xlsx_file:
        write_xlsx(nested, cond_order, xlsx_path)
        if not os.path.exists(xlsx_path):
            xlsx_path = None

    # ---------- significance testing (optional) ----------
    if do_stats:
        logger.info("  harvesting per-replicate scalars for stats")
        per_rep, per_rep_by_rep, reps_per_cond = \
            harvest_per_replicate_scalars(analysis_dir)
        # Keep only conditions that actually have per-rep scalars
        stats_conds = [c for c in cond_order if c in reps_per_cond]
        # Fall back to alphabetical order of whatever per-rep conds exist
        # if none of them match cond_order (unexpected, but defensive).
        if not stats_conds:
            stats_conds = sorted(reps_per_cond.keys())

        # Check that at least two conditions have >= min_n replicates for
        # at least one shared metric; otherwise skip stats.
        usable = False
        for group_metrics in per_rep.values():
            for cond_map in group_metrics.values():
                qualifying = [c for c, vs in cond_map.items()
                              if len(vs) >= min_n_per_condition]
                if len(qualifying) >= 2:
                    usable = True
                    break
            if usable:
                break

        if not usable:
            logger.info(
                f"  no metric has two conditions each with "
                f">= {min_n_per_condition} replicate values; "
                f"skipping stats CSVs.")
        else:
            stats_dir = os.path.dirname(os.path.abspath(out_prefix))

            # All-ordered-pairs table
            stats_all, pair_labels_all = run_pairwise_stats(
                per_rep, stats_conds, pair_mode="all_ordered",
                n_boot=n_boot, seed=stats_seed,
                min_n_per_condition=min_n_per_condition,
            )
            if pair_labels_all:
                write_stats_csv(stats_all, pair_labels_all,
                                os.path.join(stats_dir,
                                             "summary_stats_all_pairs.csv"))
                write_stats_json(
                    stats_all, pair_labels_all,
                    pair_mode="all_ordered", reference=None,
                    n_boot=n_boot,
                    out_path=os.path.join(
                        stats_dir, "summary_stats_all_pairs.json"),
                )

            # vs-reference table
            ref = reference_condition or stats_conds[0]
            if ref not in stats_conds:
                logger.warning(
                    f"  reference condition {ref!r} not in per-rep "
                    f"conditions {stats_conds!r}; falling back to "
                    f"{stats_conds[0]!r}.")
                ref = stats_conds[0]
            stats_ref, pair_labels_ref = run_pairwise_stats(
                per_rep, stats_conds, pair_mode="vs_reference",
                reference=ref,
                n_boot=n_boot, seed=stats_seed,
                min_n_per_condition=min_n_per_condition,
            )
            if pair_labels_ref:
                write_stats_csv(
                    stats_ref, pair_labels_ref,
                    os.path.join(stats_dir, "summary_stats_vs_reference.csv"))
                write_stats_json(
                    stats_ref, pair_labels_ref,
                    pair_mode="vs_reference", reference=ref,
                    n_boot=n_boot,
                    out_path=os.path.join(
                        stats_dir, "summary_stats_vs_reference.json"),
                )

            # Per-replicate audit
            write_per_replicate_csv(
                per_rep_by_rep, reps_per_cond,
                os.path.join(stats_dir, "summary_stats_per_replicate.csv"),
            )

    return csv_path, json_path, xlsx_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Collect all per-condition analysis results into a single "
            "metric-by-condition summary table. See cohesin_sim/README.md "
            "section 8 for what each row means."
        ),
    )
    parser.add_argument(
        "--analysis-dir", type=str, required=True,
        help="Directory with pooled outputs (the --output-dir you passed "
             "to run_analysis_all.py).",
    )
    parser.add_argument(
        "--out-prefix", type=str, default=None,
        help="Prefix for output files (no extension). Default: "
             "<analysis-dir>/summary_analysis_table",
    )
    parser.add_argument(
        "--no-xlsx", action="store_true",
        help="Skip the XLSX output even if openpyxl is available.",
    )
    parser.add_argument(
        "--no-stats-in-summary", action="store_true",
        help="Skip the significance-testing companion CSVs "
             "(summary_stats_all_pairs.csv / summary_stats_vs_reference.csv).",
    )
    parser.add_argument(
        "--reference-condition", type=str, default=None,
        help="Condition to use as the reference in "
             "summary_stats_vs_reference.csv. Defaults to the first "
             "condition discovered.",
    )
    parser.add_argument(
        "--n-boot", type=int, default=10_000,
        help="Bootstrap / permutation iterations for the stats. "
             "Default 10000.",
    )
    parser.add_argument(
        "--stats-seed", type=int, default=0,
        help="Seed for bootstrap / permutation RNGs. Default 0.",
    )
    parser.add_argument(
        "--min-n-per-condition", type=int, default=2,
        help="Minimum per-replicate values a condition must contribute "
             "to a metric to be included in a pairwise test. Default 2.",
    )
    args = parser.parse_args()

    if not os.path.isdir(args.analysis_dir):
        logger.error(f"Not a directory: {args.analysis_dir}")
        sys.exit(1)

    csv_path, json_path, xlsx_path = build_summary_table(
        args.analysis_dir,
        out_prefix=args.out_prefix,
        write_xlsx_file=not args.no_xlsx,
        do_stats=not args.no_stats_in_summary,
        reference_condition=args.reference_condition,
        n_boot=args.n_boot,
        stats_seed=args.stats_seed,
        min_n_per_condition=args.min_n_per_condition,
    )
    if not csv_path:
        sys.exit(1)
    logger.info("Summary table written.")


if __name__ == "__main__":
    main()
